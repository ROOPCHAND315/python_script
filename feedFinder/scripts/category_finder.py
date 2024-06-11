
from deep_translator import GoogleTranslator
from rich import print
from urllib.parse import urlparse
import openpyxl
import feedparser
from langdetect import detect
def category_finder():
    category_dict = {
        'bus': 'business',
        'ent': 'entertainment',
        'env': 'environment',
        'foo': 'food',
        'hea': 'health',
        'pol': 'politics',
        'sci': 'science',
        'spo': 'sports',
        'tec': 'technology',
        'top': 'top',
        'tou': 'tourism',
        'wor': 'world',
        'cri':'crime',
        'dom':'domestic',
        'edu':'education',
        'oth':'other',
        'lif':'lifestyle'
    }
    short_keys = list(category_dict.keys())
    import pandas as pd
    df=pd.read_excel('/home/roopchand/test/scrapyTest/feedFinder/feedFinder/scripts/sheets/collect_data_sheet.xlsx',sheet_name='Data')
    data=df['source'].to_list()
    # base=df.loc[:0,'Domain'].to_list()

    category_dict2 = {
        'business': ['business','company','companies','economic','economy','finance', 'market','shop','startups','job'],
        'entertainment': ['entertainment','showbiz','shows','film','cultural','culture','arts','wedding','movies', 'gaming','celebrities','television','web-series','hollywood','relationships','bollywood','festival','/tv/','cinema','photo'],
        'environment': ['environment','climate', 'nature','earth','weather'],
        'food': ['food','restaurants', 'recipes','gastronomy'],
        'health': ['health','wellness', 'fitness','coronavirus'],
        'politics': ['politics','policiales','government', 'elections','security','justice'],
        'science': ['science','research', 'discovery','physics','space','astrology'],
        'sports': ['sports','sport','athletics', 'competitions','golf', 'football', 'basketball','/ipl', 'tennis', 'cricket', 'baseball', 'rugby', 'hockey'],
        'technology': ['technology','automobiles','gadgets', 'innovation','tech','smartphones','software','tablets'],
        'top': ['top/','local','things-to-do/','famous','society','popular', 'trending','feature','breaking','magazine','latest/','states','/national/','columns','interviews','nation'],
        'tourism': ['tourism','travel','destinations', 'adventures'],
        'world': ['world','international','global'],
        'crime':['crime','police','court'],
        'domestic':['domestic'],
        'education':['education','books'],
        'other':['other','obituaries','opinion','regional','social','vlog','people','archive','editor','edition'],
        'lifestyle':['lifestyle','life','life-style','fashion','shopping']
    }
     # Default category if keyword doesn't match any category
    def parse_feed_for_analysis(url):
        """
        Parse the RSS feed and return data for a specified number of items as a dictionary.
        """
        feed = feedparser.parse(url)
        num_items_to_parse = 2
        parsed_items = 0
        items = {}

        for entry in feed.entries:
            title = entry.title
            description = entry.description if 'description' in entry else 'Description not found' 

            if hasattr(entry, 'tags') and entry.tags:
                categories = [category['term'] for category in entry.tags]
            else:
                categories = []
            item_data = {
                "title": title,
                "description": description,
                "categories": categories
            }

            # Use item ID as the key in the dictionary
            item_id = entry.id
            items[item_id] = item_data

            parsed_items += 1
            if parsed_items >= num_items_to_parse:
                break
            
        return items


    def language_detect(data:str):
        """
        Detect the language of the input text.
        """
        try:
            # Check if the input text is not empty
            if data.strip():
                language = detect(data)
            else:
                language = "Unknown"
        except Exception as e:
            print(f"An error occurred: {e}")
            language = "Unknown"
        return language

    def translate_if_not_english(text, target_language='en'):
        """
        Translates the text to English if the detected language is not English.
        """
        detected_language = language_detect(text)
        # print(f"detect lan {detected_language}")
        if detected_language != 'en' and text.strip():  # Check if text is not empty
            try:
                translated_text = GoogleTranslator(source=detected_language, target=target_language).translate(text)
                return translated_text
            except Exception as e:
                print(f"An error occurred during translation: {e}")
        return text  


    def find_category(url):
        for category, keywords_list in category_dict2.items():
            if isinstance(keywords_list, list):
                for cat in keywords_list:
                    if cat in url:
                        return(category)

    category_list = []
            
    # print(f"{base[0]} {type(base[0])}")
    for url_link in data:
        parsed_url = urlparse(url_link)
        path = (parsed_url.path).replace('feed','').replace('.xml','').replace('.','')
        lst=path.split('/')
        other_lang=[]
        for part_url in lst:
            if part_url is not None and part_url != '':
                other_lang.append(translate_if_not_english(part_url))
        other_lang=','.join(other_lang)
        print(f"splited link {other_lang}")
        url=other_lang
        if find_category(url):
            category_list.append(find_category(url))
            print(f"[red] Auto added {find_category(url)} for {lst} --> {url}")
        else:
            def transelate():
                try:
                    
                    parsed_items = parse_feed_for_analysis(url_link)
                    for item_id, item_data in parsed_items.items():
                        print("Item ID:", item_id)
                        title = translate_if_not_english(str(item_data['title']))
                        print("Title:", title)
                        if item_data['description']:
                            description = translate_if_not_english(str(item_data['description']))
                            print("Description:", description)
                        if item_data['categories']:   
                            category=','.join(item_data['categories'])
                            print("Categories:",translate_if_not_english(str(category) ))
                        print("="*50)
                  # Exit the loop if input is successfully obtained
                except ValueError:
                    print("Invalid input! Please enter 1 to see translated title, description, and category. else any number")

            print(f"[yellow] Enter the category key for URL:  {other_lang} -->  {url_link} ")  
            user_input = input().strip().lower()
            if user_input=='tra':
                transelate()
                user_input = input().strip().lower()
            while user_input not in short_keys:
                print("Invalid category key! Please enter a valid category key from the list:")
                print(short_keys)
                user_input = input().strip().lower()
            try:    
            # Append the corresponding category value to the category_list
                category_list.append(category_dict[user_input])
            except:
                print(f"No Key match!")    

            print(f"added {category_dict[user_input]}")
    
        print("Category list:", category_list)
        wb = openpyxl.load_workbook('/home/roopchand/test/scrapyTest/feedFinder/feedFinder/scripts/sheets/collect_data_sheet.xlsx')
        # if 'url1' not in wb.sheetnames:
        #     wb.create_sheet('url1')
        sheet = wb['Data']
        # Add column name
        sheet['I1'] = "Category"
        for i, url in enumerate(category_list, start=2):  # Start from row 2 to leave the first row for the column name
            sheet.cell(row=i, column=9, value=url)
        wb.save('/home/roopchand/test/scrapyTest/feedFinder/feedFinder/scripts/sheets/collect_data_sheet.xlsx')
        print(f"save category succesfully !!")
# category_finder()